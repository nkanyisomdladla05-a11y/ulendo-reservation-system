import os

from django.conf import settings
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import date


def generate_pdf_report(report_type, report_date, data):
    """Generate PDF report using reportlab."""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{report_date}.pdf"'
    
    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch,
        topMargin=0.15*inch,
        bottomMargin=0.75*inch,
    )
    elements = []
    styles = getSampleStyleSheet()

    # Header: logo + contact info, top center
    contact_lines = [
        '<font color="#D4AF37" size="16">★ ★ ★ ★</font>',
        '<b><font size="18">Ulendo Lodge & Apartments</font></b>',
        '<font size="14" color="#444444">10 Sinclair Road, Lambton, Germiston, 1401</font>',
        '<font size="13" color="#555555">Tel: 067 623 7170 &nbsp;&nbsp; Tel: 010 824 4595</font>',
        '<font size="13" color="#555555">Email: info@ulendolodge.com</font>',
        '<font size="13" color="#666666">Reg Nr. 2016/078946/07</font>',
    ]
    contact_text = '<br/>'.join(contact_lines)
    contact_style = ParagraphStyle(
        'Contact',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        alignment=1,  # TA_CENTER
        leftIndent=0,
        rightIndent=0,
        spaceBefore=0,
        spaceAfter=2,
        leading=19,
    )
    logo_path = os.path.join(settings.BASE_DIR, 'assets', 'logo.png')
    logo_img = None
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=3.0 * inch, height=3.0 * inch)
            logo_img.hAlign = 'CENTER'
        except Exception:
            pass
    contact_para = Paragraph(contact_text, contact_style)
    if logo_img:
        header_data = [[logo_img], [Spacer(1, 0.08*inch)], [contact_para]]
        header_table = Table(header_data, colWidths=[5.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
    else:
        header_data = [[contact_para]]
        header_table = Table(header_data, colWidths=[5.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.25 * inch))

    # Thin divider line
    line_table = Table([['']], colWidths=[6.5*inch])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#DDDDDD')),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2 * inch))

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#C77A1A'),
        spaceAfter=30,
    )
    # For daily-style reports we may have range modes (daily/weekly/monthly/custom)
    if report_type == 'daily':
        mode = data.get('mode', 'daily')
        mode_label_map = {
            'daily': 'Daily',
            'weekly': 'Weekly',
            'monthly': 'Monthly',
            'custom': 'Custom Range',
        }
        mode_label = mode_label_map.get(mode, 'Daily')
        title_text = f'{mode_label} Report - {report_date}'
    elif report_type == 'monthly':
        # Monthly exports still use the month as the title date
        if isinstance(report_date, date):
            title_text = f'Monthly Report - {report_date.strftime("%B %Y")}'
        else:
            title_text = f'Monthly Report - {report_date}'
    else:
        title_text = f'{report_type.title()} Report - {report_date}'

    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.2*inch))
    
    if report_type == 'daily':
        # Daily/weekly/monthly/custom (range) report content
        range_label = data.get('date_label', str(report_date))
        mode = data.get('mode', 'daily')
        mode_label_map = {
            'daily': 'Daily',
            'weekly': 'Weekly',
            'monthly': 'Monthly',
            'custom': 'Custom Range',
        }
        mode_label = mode_label_map.get(mode, 'Daily')

        elements.append(Paragraph(f'<b>Report type:</b> {mode_label}', styles['Normal']))
        elements.append(Paragraph(f'<b>Period:</b> {range_label}', styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Single Booking Information table (combines check-ins and check-outs, deduplicated)
        seen_ids = set()
        all_reservations = []
        for r in list(data['check_ins']) + list(data['check_outs']):
            if r.id not in seen_ids:
                seen_ids.add(r.id)
                all_reservations.append(r)
        all_reservations.sort(key=lambda r: (r.room.room_number, r.check_in_date))

        if all_reservations:
            elements.append(Paragraph('<b>Booking Information</b>', styles['Heading2']))
            booking_data = [['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']]
            for reservation in all_reservations:
                booking_data.append([
                    f"Room {reservation.room.room_number}",
                    reservation.voucher_number or '-',
                    reservation.customer_name,
                    reservation.check_in_date.strftime('%Y-%m-%d'),
                    reservation.check_out_date.strftime('%Y-%m-%d')
                ])
            col_widths = [1.2*inch, 1.8*inch, 2.4*inch, 1.4*inch, 1.4*inch]
            table = Table(booking_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F0F0F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Summary
        summary_data = [
            ['Total Rooms', str(data['total_rooms'])],
            ['Booked Rooms', str(data['booked_rooms'])],
            ['Occupancy Rate', f"{data['occupancy_rate']:.1f}%"],
        ]
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C77A1A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
    
    elif report_type == 'monthly':
        # Monthly report content
        elements.append(Paragraph(f'<b>Month:</b> {report_date.strftime("%B %Y")}', styles['Normal']))
        elements.append(Spacer(1, 0.2*inch))
        
        # Summary
        summary_data = [
            ['Total Bookings', str(data['total_bookings'])],
            ['Average Occupancy', f"{data['avg_occupancy']:.1f}%"],
            ['Total Rooms', str(data['total_rooms'])],
        ]
        table = Table(summary_data, colWidths=[3*inch, 2*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C77A1A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)

        # Reservations detail table
        if data.get('reservations'):
            elements.append(Spacer(1, 0.2*inch))
            elements.append(Paragraph('<b>Booking Information</b>', styles['Heading2']))
            reserv_data = [['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']]
            for r in data['reservations']:
                reserv_data.append([
                    f"Room {r.room.room_number}",
                    r.voucher_number or '-',
                    r.customer_name,
                    r.check_in_date.strftime('%Y-%m-%d'),
                    r.check_out_date.strftime('%Y-%m-%d')
                ])
            col_widths = [1.2*inch, 1.8*inch, 2.4*inch, 1.4*inch, 1.4*inch]
            table = Table(reserv_data, colWidths=col_widths)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F0F0F')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(table)
    
    doc.build(elements)
    return response


def generate_excel_report(report_type, report_date, data):
    """Generate Excel report using openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = f'{report_type.title()} Report'
    
    # Header style
    header_fill = PatternFill(start_color='C77A1A', end_color='C77A1A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)
    
    # Title
    ws['A1'] = f'{report_type.title()} Report - {report_date}'
    ws['A1'].font = Font(bold=True, size=14, color='C77A1A')
    ws.merge_cells('A1:D1')
    
    row = 3
    
    if report_type == 'daily':
        # Check-ins
        ws[f'A{row}'] = 'Check-ins'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for reservation in data['check_ins']:
            ws.cell(row=row, column=1, value=f"Room {reservation.room.room_number}")
            ws.cell(row=row, column=2, value=reservation.voucher_number or '-')
            ws.cell(row=row, column=3, value=reservation.customer_name)
            ws.cell(row=row, column=4, value=reservation.check_in_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=reservation.check_out_date.strftime('%Y-%m-%d'))
            row += 1
        
        row += 2
        
        # Check-outs
        ws[f'A{row}'] = 'Check-outs'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        headers = ['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        row += 1
        for reservation in data['check_outs']:
            ws.cell(row=row, column=1, value=f"Room {reservation.room.room_number}")
            ws.cell(row=row, column=2, value=reservation.voucher_number or '-')
            ws.cell(row=row, column=3, value=reservation.customer_name)
            ws.cell(row=row, column=4, value=reservation.check_in_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=5, value=reservation.check_out_date.strftime('%Y-%m-%d'))
            row += 1
        
        row += 2
        
        # Summary
        ws[f'A{row}'] = 'Summary'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        summary_data = [
            ['Total Rooms', data['total_rooms']],
            ['Booked Rooms', data['booked_rooms']],
            ['Occupancy Rate', f"{data['occupancy_rate']:.1f}%"],
        ]
        
        for item in summary_data:
            ws.cell(row=row, column=1, value=item[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=item[1])
            row += 1
    
    elif report_type == 'monthly':
        # Summary
        ws['A3'] = 'Summary'
        ws['A3'].font = Font(bold=True, size=12)
        
        summary_data = [
            ['Total Bookings', data['total_bookings']],
            ['Average Occupancy', f"{data['avg_occupancy']:.1f}%"],
            ['Total Rooms', data['total_rooms']],
        ]
        
        row = 4
        for item in summary_data:
            ws.cell(row=row, column=1, value=item[0]).font = Font(bold=True)
            ws.cell(row=row, column=2, value=item[1])
            row += 1

        # Reservations detail table
        if data.get('reservations'):
            row += 2
            ws.cell(row=row, column=1, value='Reservations').font = Font(bold=True, size=12)
            row += 1
            headers = ['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
            row += 1
            for r in data['reservations']:
                ws.cell(row=row, column=1, value=f"Room {r.room.room_number}")
                ws.cell(row=row, column=2, value=r.voucher_number or '-')
                ws.cell(row=row, column=3, value=r.customer_name)
                ws.cell(row=row, column=4, value=r.check_in_date.strftime('%Y-%m-%d'))
                ws.cell(row=row, column=5, value=r.check_out_date.strftime('%Y-%m-%d'))
                row += 1
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{report_type}_report_{report_date}.xlsx"'
    wb.save(response)
    return response
