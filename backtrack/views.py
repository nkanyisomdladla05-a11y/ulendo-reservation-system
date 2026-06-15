import logging
import re

from django.db import transaction
from django.db.models import IntegerField
from django.db.models.functions import Cast
from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.contrib import messages
from django.core.paginator import Paginator
from django.http import HttpResponse
from datetime import date, datetime, timedelta
import calendar
from dateutil import parser as date_parser

from .models import BacktrackReservation, BacktrackVoucher
from rooms.models import Room
from .forms import BacktrackReservationForm, BacktrackVoucherUploadForm
from .services import create_backtrack_reservation
from vouchers.services import extract_voucher_data

logger = logging.getLogger(__name__)


def parse_date_safe(date_str):
    """Parse date string safely, handling YYYY/MM/DD and YY/MM/DD formats correctly."""
    if not date_str:
        return None
    date_str = date_str.strip()
    ymd = re.match(r'(\d{4})[/-](\d{1,2})[/-](\d{1,2})', date_str)
    if ymd:
        try:
            return date(int(ymd.group(1)), int(ymd.group(2)), int(ymd.group(3)))
        except ValueError:
            return None
    dmy = re.match(r'(\d{1,2})[/-](\d{1,2})[/-](\d{4})', date_str)
    if dmy:
        try:
            return date(int(dmy.group(3)), int(dmy.group(2)), int(dmy.group(1)))
        except ValueError:
            return None
    ymd2 = re.match(r'(\d{2})[/-](\d{1,2})[/-](\d{2})$', date_str)
    if ymd2:
        try:
            year = int(ymd2.group(1))
            year = year + 2000 if year < 100 else year
            return date(year, int(ymd2.group(2)), int(ymd2.group(3)))
        except ValueError:
            return None
    try:
        return date_parser.parse(date_str, dayfirst=True).date()
    except ValueError:
        return None


@login_required
def dashboard(request):
    """Backtrack dashboard showing historical reservation stats."""
    today = date.today()
    total = BacktrackReservation.objects.count()
    confirmed = BacktrackReservation.objects.filter(status='confirmed').count()
    cancelled = BacktrackReservation.objects.filter(status='cancelled').count()
    recent = BacktrackReservation.objects.prefetch_related('vouchers').order_by('-created_at')[:10]

    context = {
        'total': total,
        'confirmed': confirmed,
        'cancelled': cancelled,
        'recent': recent,
        'today': today,
    }
    return render(request, 'backtrack/dashboard.html', context)


@login_required
def new_backtrack(request):
    """New backtrack reservation - manual booking for past dates."""
    if request.method == 'POST':
        check_in_str = request.POST.get('check_in_date', '').strip()
        check_out_str = request.POST.get('check_out_date', '').strip()
        room_number = request.POST.get('room_number', '').strip()

        check_in = None
        check_out = None

        if check_in_str:
            try:
                check_in = parse_date_safe(check_in_str)
            except ValueError:
                check_in = None
                messages.error(request, f'Invalid check-in date: {check_in_str}')

        if check_out_str:
            try:
                check_out = parse_date_safe(check_out_str)
            except ValueError:
                check_out = None
                messages.error(request, f'Invalid check-out date: {check_out_str}')

        dates_valid = False
        if check_in and check_out:
            if check_out > check_in:
                dates_valid = True
            else:
                messages.error(request, 'Check-out date must be after check-in date.')

        if dates_valid:
            form = BacktrackReservationForm(request.POST, check_in_date=check_in, check_out_date=check_out)
        else:
            form = BacktrackReservationForm(request.POST)

        if room_number and form.is_valid():
            cd = form.cleaned_data
            reservation = create_backtrack_reservation(
                customer_name=cd['customer_name'],
                voucher_number=cd.get('voucher_number') or '',
                confirmation_code=cd.get('confirmation_code') or '',
                room_number=cd['room_number'],
                check_in_date=cd['check_in_date'],
                check_out_date=cd['check_out_date'],
                notes=cd.get('notes') or '',
            )
            if reservation:
                messages.success(request, f'Backtrack reservation confirmed for {reservation.customer_name} in Room {reservation.room_number}.')
                return redirect('backtrack:dashboard')
            messages.error(request, 'Could not create backtrack reservation. Please check your data.')

        context = {
            'form': form,
            'check_in': check_in_str,
            'check_out': check_out_str,
        }
        return render(request, 'backtrack/new_backtrack.html', context)

    else:
        check_in_str = request.GET.get('check_in', '').strip()
        check_out_str = request.GET.get('check_out', '').strip()
        check_in = None
        check_out = None
        if check_in_str:
            try:
                check_in = parse_date_safe(check_in_str)
            except ValueError:
                check_in = None
        if check_out_str:
            try:
                check_out = parse_date_safe(check_out_str)
            except ValueError:
                check_out = None

        if not check_in or not check_out or check_out <= check_in:
            check_in = date.today()
            check_out = date.today() + timedelta(days=1)
            check_in_str = check_in.strftime('%Y/%m/%d')
            check_out_str = check_out.strftime('%Y/%m/%d')

        form = BacktrackReservationForm(check_in_date=check_in, check_out_date=check_out)

        return render(request, 'backtrack/new_backtrack.html', {
            'form': form,
            'check_in': check_in_str,
            'check_out': check_out_str,
        })


@login_required
def upload_backtrack_voucher(request):
    """Upload voucher for backtrack processing."""
    if request.method == 'POST':
        form = BacktrackVoucherUploadForm(request.POST, request.FILES)
        manual_check_in = request.POST.get('check_in_date', '').strip()
        manual_check_out = request.POST.get('check_out_date', '').strip()
        if form.is_valid():
            voucher = form.save()
            try:
                extracted = extract_voucher_data(voucher.voucher_file.path)

                extracted_json = extracted.copy() if isinstance(extracted, dict) else {}
                ci = extracted_json.get('check_in_date')
                co = extracted_json.get('check_out_date')
                if isinstance(ci, date):
                    extracted_json['check_in_date'] = ci.strftime('%Y/%m/%d')
                if isinstance(co, date):
                    extracted_json['check_out_date'] = co.strftime('%Y/%m/%d')

                voucher.extracted_data = extracted_json
                voucher.customer_name = extracted.get('customer_name', '')
                voucher.voucher_number = extracted.get('voucher_number', '')
                voucher.check_in_date = extracted.get('check_in_date')
                voucher.check_out_date = extracted.get('check_out_date')
                voucher.check_in_raw = extracted.get('check_in_raw', '')
                voucher.check_out_raw = extracted.get('check_out_raw', '')
                voucher.save(update_fields=[
                    'extracted_data', 'customer_name', 'voucher_number',
                    'check_in_date', 'check_out_date', 'check_in_raw', 'check_out_raw'
                ])

                # Debug log
                import sys
                print(f"\n=== BACKTRACK VOUCHER SAVED ===", file=sys.stderr)
                print(f"Saved check_in: {voucher.check_in_date}", file=sys.stderr)
                print(f"Saved check_out: {voucher.check_out_date}", file=sys.stderr)
                print(f"Saved check_in_raw: {voucher.check_in_raw}", file=sys.stderr)
                print(f"Saved check_out_raw: {voucher.check_out_raw}", file=sys.stderr)
                print(f"==============================\n", file=sys.stderr)

                review_url = reverse('backtrack:review_backtrack_voucher', kwargs={'voucher_id': voucher.id})
                params = []
                if manual_check_in:
                    params.append(f'check_in={manual_check_in}')
                if manual_check_out:
                    params.append(f'check_out={manual_check_out}')
                if params:
                    review_url += '?' + '&'.join(params)
                return redirect(review_url)
            except Exception as e:
                voucher.delete()
                messages.error(request, f'Error processing voucher: {str(e)}')
                return render(request, 'backtrack/new_backtrack.html', {'upload_form': form})
    else:
        form = BacktrackVoucherUploadForm()

    return render(request, 'backtrack/new_backtrack.html', {'upload_form': form})


@login_required
def review_backtrack_voucher(request, voucher_id):
    """Review OCR extracted data for backtrack voucher and confirm reservation."""
    voucher = get_object_or_404(BacktrackVoucher, pk=voucher_id)

    if request.method == 'POST':
        if getattr(voucher, 'reservation_id', None) and voucher.is_confirmed:
            messages.info(request, 'This voucher is already confirmed.')
            return redirect('backtrack:dashboard')

        voucher.customer_name = request.POST.get('customer_name', '')
        voucher.voucher_number = request.POST.get('voucher_number', '')

        check_in_str = request.POST.get('check_in_date', '').strip()
        check_out_str = request.POST.get('check_out_date', '').strip()
        voucher.check_in_date = None
        voucher.check_out_date = None
        if check_in_str:
            try:
                voucher.check_in_date = parse_date_safe(check_in_str)
            except ValueError:
                pass
        if check_out_str:
            try:
                voucher.check_out_date = parse_date_safe(check_out_str)
            except ValueError:
                pass
        voucher.save()

        has_valid_dates = (
            isinstance(voucher.check_in_date, date)
            and isinstance(voucher.check_out_date, date)
            and voucher.check_out_date > voucher.check_in_date
        )

        room_form = BacktrackReservationForm(
            request.POST,
            initial={
                'customer_name': voucher.customer_name,
                'voucher_number': voucher.voucher_number,
                'check_in_date': check_in_str,
                'check_out_date': check_out_str,
            },
            check_in_date=voucher.check_in_date if has_valid_dates else None,
            check_out_date=voucher.check_out_date if has_valid_dates else None,
        )

        room_number = request.POST.get('room_number')
        if room_number and has_valid_dates and room_form.is_valid():
            reservation = create_backtrack_reservation(
                customer_name=voucher.customer_name,
                voucher_number=voucher.voucher_number or '',
                confirmation_code=request.POST.get('confirmation_code', '') or '',
                room_number=room_number,
                check_in_date=voucher.check_in_date,
                check_out_date=voucher.check_out_date,
                notes='',
            )
            if reservation:
                def link_voucher():
                    try:
                        voucher.reservation = reservation
                        voucher.is_confirmed = True
                        voucher.save(update_fields=['reservation', 'is_confirmed'])
                    except Exception:
                        logger.exception("Failed to link backtrack voucher to reservation")
                transaction.on_commit(link_voucher)
                messages.success(request, f'Backtrack reservation confirmed for {reservation.customer_name} in Room {reservation.room_number}.')
                return redirect('backtrack:backtrack_list')
            messages.error(request, 'Could not create backtrack reservation. Please check your data.')
        elif not has_valid_dates and (check_in_str or check_out_str):
            messages.error(request, 'Please set valid check-in and check-out dates.')
        else:
            if room_form.errors:
                for field, errors in room_form.errors.items():
                    for error in errors:
                        messages.error(request, error)
            else:
                messages.error(request, 'Please select a room and ensure dates are set.')

    form_check_in = voucher.check_in_date if isinstance(voucher.check_in_date, date) else None
    form_check_out = voucher.check_out_date if isinstance(voucher.check_out_date, date) else None

    # Filter room dropdown by current availability (today's date), like manual booking
    today = date.today()
    tomorrow = today + timedelta(days=1)
    room_form = BacktrackReservationForm(
        initial={
            'customer_name': voucher.customer_name,
            'voucher_number': voucher.voucher_number,
            'check_in_date': form_check_in.strftime('%Y/%m/%d') if form_check_in else '',
            'check_out_date': form_check_out.strftime('%Y/%m/%d') if form_check_out else '',
        },
        check_in_date=today,
        check_out_date=tomorrow,
    )

    context = {
        'voucher': voucher,
        'room_form': room_form,
        'check_in': form_check_in.strftime('%Y/%m/%d') if form_check_in else '',
        'check_out': form_check_out.strftime('%Y/%m/%d') if form_check_out else '',
    }

    # Debug log
    import sys
    print(f"\n=== BACKTRACK REVIEW PAGE ===", file=sys.stderr)
    print(f"form_check_in: {form_check_in} form_check_out: {form_check_out}", file=sys.stderr)
    print(f"context check_in: '{context['check_in']}' check_out: '{context['check_out']}'", file=sys.stderr)
    print(f"voucher.check_in_date: {voucher.check_in_date}", file=sys.stderr)
    print(f"voucher.check_out_date: {voucher.check_out_date}", file=sys.stderr)
    print(f"voucher.extracted_data: {dict(voucher.extracted_data) if voucher.extracted_data else None}", file=sys.stderr)
    print(f"============================\n", file=sys.stderr)

    return render(request, 'backtrack/review_backtrack.html', context)


@login_required
def backtrack_list(request):
    """List all backtrack reservations with search."""
    reservations = BacktrackReservation.objects.prefetch_related('vouchers').order_by('-check_in_date')
    search_name = request.GET.get('name', '')
    search_voucher = request.GET.get('voucher', '')
    start_date = request.GET.get('start_date', '')
    end_date = request.GET.get('end_date', '')

    if search_name:
        reservations = reservations.filter(customer_name__icontains=search_name)
    if search_voucher:
        reservations = reservations.filter(voucher_number__icontains=search_voucher)
    if start_date:
        try:
            start_date_obj = parse_date_safe(start_date)
            if start_date_obj:
                reservations = reservations.filter(check_in_date__gte=start_date_obj)
        except ValueError:
            pass
    if end_date:
        try:
            end_date_obj = parse_date_safe(end_date)
            if end_date_obj:
                reservations = reservations.filter(check_out_date__lte=end_date_obj)
        except ValueError:
            pass

    paginator = Paginator(reservations, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'search_name': search_name,
        'search_voucher': search_voucher,
        'start_date': start_date,
        'end_date': end_date,
    }
    return render(request, 'backtrack/backtrack_list.html', context)


@login_required
def backtrack_report(request):
    """Generate reports for backtrack reservations."""
    mode = request.GET.get('mode', 'daily')

    # Default date: use latest backtrack reservation check_in_date, not today
    latest = BacktrackReservation.objects.order_by('-check_in_date').first()
    default_date = latest.check_in_date if latest else date.today()
    date_str = request.GET.get('date', default_date.strftime('%y/%m/%d'))
    try:
        anchor_date = parse_date_safe(date_str) or default_date
    except ValueError:
        anchor_date = default_date

    if mode == 'weekly':
        start_date = anchor_date - timedelta(days=anchor_date.weekday())
        end_date = start_date + timedelta(days=6)
    elif mode == 'monthly':
        start_date = anchor_date.replace(day=1)
        if start_date.month == 12:
            next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
        else:
            next_month = start_date.replace(month=start_date.month + 1, day=1)
        end_date = next_month - timedelta(days=1)
    elif mode == 'custom':
        start_str = request.GET.get('start_date')
        end_str = request.GET.get('end_date')
        try:
            start_date = parse_date_safe(start_str) if start_str else anchor_date
        except ValueError:
            start_date = anchor_date
        try:
            end_date = parse_date_safe(end_str) if end_str else start_date
        except ValueError:
            end_date = start_date
        if end_date < start_date:
            end_date = start_date
    else:
        mode = 'daily'
        start_date = end_date = anchor_date

    check_ins = BacktrackReservation.objects.filter(
        check_in_date__gte=start_date,
        check_in_date__lte=end_date,
        status='confirmed'
    ).order_by('room_number')

    check_outs = BacktrackReservation.objects.filter(
        check_out_date__gte=start_date,
        check_out_date__lte=end_date,
        status='confirmed'
    ).order_by('room_number')

    total = BacktrackReservation.objects.filter(
        check_in_date__lt=end_date + timedelta(days=1),
        check_out_date__gt=start_date,
        status='confirmed'
    ).values('room_number').distinct().count()

    if start_date == end_date:
        date_label = start_date.strftime('%y/%m/%d')
    else:
        date_label = f"{start_date.strftime('%y/%m/%d')} – {end_date.strftime('%y/%m/%d')}"

    export_format = request.GET.get('export')
    if export_format == 'pdf':
        return generate_backtrack_pdf_report(date_label, check_ins, check_outs, total, mode)
    elif export_format == 'excel':
        return generate_backtrack_excel_report(date_label, check_ins, check_outs, total, mode)

    context = {
        'mode': mode,
        'anchor_date': anchor_date,
        'start_date': start_date,
        'end_date': end_date,
        'date_label': date_label,
        'check_ins': check_ins,
        'check_outs': check_outs,
        'total': total,
    }
    return render(request, 'backtrack/reports.html', context)


@login_required
def edit_backtrack_reservation(request, pk):
    """Edit an existing backtrack reservation."""
    reservation = get_object_or_404(BacktrackReservation, pk=pk)
    if request.method == 'POST':
        form = BacktrackReservationForm(request.POST, instance=reservation)
        if form.is_valid():
            form.save()
            messages.success(request, f'Backtrack reservation for {reservation.customer_name} has been updated.')
            return redirect('backtrack:backtrack_list')
    else:
        form = BacktrackReservationForm(instance=reservation)
    return render(request, 'backtrack/edit_backtrack.html', {'form': form, 'reservation': reservation})


@login_required
def delete_backtrack_reservation(request, pk):
    """Delete a backtrack reservation."""
    reservation = get_object_or_404(BacktrackReservation, pk=pk)
    if request.method == 'POST':
        reservation.delete()
        messages.success(request, f'Backtrack reservation for {reservation.customer_name} has been deleted.')
    return redirect('backtrack:backtrack_list')


@login_required
def delete_backtrack_from_dashboard(request, pk):
    """Delete a backtrack reservation from dashboard."""
    reservation = get_object_or_404(BacktrackReservation, pk=pk)
    if request.method == 'POST':
        reservation.delete()
        messages.success(request, f'Backtrack reservation for {reservation.customer_name} has been deleted.')
    return redirect('backtrack:dashboard')


@login_required
def api_available_backtrack_rooms(request):
    """AJAX endpoint: return available rooms as JSON for given dates (backtrack system)."""
    from .forms import get_available_backtrack_rooms
    check_in_str = request.GET.get('check_in', '').strip()
    check_out_str = request.GET.get('check_out', '').strip()
    if not check_in_str or not check_out_str:
        return JsonResponse({'rooms': [], 'count': 0})
    try:
        check_in = parse_date_safe(check_in_str)
        check_out = parse_date_safe(check_out_str)
    except Exception:
        return JsonResponse({'rooms': [], 'count': 0})
    if not check_in or not check_out or check_out <= check_in:
        return JsonResponse({'rooms': [], 'count': 0})
    available = get_available_backtrack_rooms(check_in, check_out)
    rooms = [{'id': r.room_number, 'label': str(r)} for r in available]
    return JsonResponse({'rooms': rooms, 'count': len(rooms)})


def generate_backtrack_pdf_report(date_label, check_ins, check_outs, total, mode):
    """Generate PDF report for backtrack reservations using room_number string field."""
    import os
    from django.conf import settings
    from django.http import HttpResponse
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="backtrack_report_{date_label.replace("/", "-")}.pdf"'

    doc = SimpleDocTemplate(
        response,
        pagesize=letter,
        leftMargin=0.75*inch,
        rightMargin=0.75*inch,
        topMargin=0.15*inch,
        bottomMargin=0.75*inch,
    )
    elements = []
    styles = getSampleStyleSheet()

    contact_lines = [
        '<font color="#D4AF37" size="16">★ ★ ★ ★</font>',
        '<b><font size="18">Ulendo Lodge & Apartments</font></b>',
        '<font size="14" color="#444444">10 Sinclair Road, Lambton, Germiston, 1401</font>',
        '<font size="13" color="#555555">Tel: 067 623 7170 &nbsp;&nbsp; Tel: 010 824 4595</font>',
        '<font size="13" color="#555555">Email: info@ulendolodge.com</font>',
        '<font size="13" color="#666666">Reg Nr. 2016/078946/07</font>',
    ]
    contact_text = '<br/>'.join(contact_lines)
    contact_style = ParagraphStyle(
        'Contact',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#333333'),
        alignment=1,
        leftIndent=0,
        rightIndent=0,
        spaceBefore=0,
        spaceAfter=2,
        leading=19,
    )
    logo_path = os.path.join(settings.BASE_DIR, 'assets', 'logo.png')
    logo_img = None
    if os.path.exists(logo_path):
        try:
            logo_img = Image(logo_path, width=3.0 * inch, height=3.0 * inch)
            logo_img.hAlign = 'CENTER'
        except Exception:
            pass
    contact_para = Paragraph(contact_text, contact_style)
    if logo_img:
        header_data = [[logo_img], [Spacer(1, 0.08*inch)], [contact_para]]
        header_table = Table(header_data, colWidths=[5.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
        ]))
    else:
        header_data = [[contact_para]]
        header_table = Table(header_data, colWidths=[5.5*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.25 * inch))

    line_table = Table([['']], colWidths=[6.5*inch])
    line_table.setStyle(TableStyle([
        ('LINEBELOW', (0, 0), (-1, 0), 0.5, colors.HexColor('#DDDDDD')),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(line_table)
    elements.append(Spacer(1, 0.2 * inch))

    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#C77A1A'),
        spaceAfter=30,
    )
    mode_label_map = {'daily': 'Daily', 'weekly': 'Weekly', 'monthly': 'Monthly', 'custom': 'Custom Range'}
    mode_label = mode_label_map.get(mode, 'Daily')
    title_text = f'Backtrack {mode_label} Report - {date_label}'
    elements.append(Paragraph(title_text, title_style))
    elements.append(Spacer(1, 0.2*inch))

    elements.append(Paragraph(f'<b>Report type:</b> {mode_label}', styles['Normal']))
    elements.append(Paragraph(f'<b>Period:</b> {date_label}', styles['Normal']))
    elements.append(Spacer(1, 0.2*inch))

    seen_ids = set()
    all_reservations = []
    for r in list(check_ins) + list(check_outs):
        if r.id not in seen_ids:
            seen_ids.add(r.id)
            all_reservations.append(r)
    all_reservations.sort(key=lambda r: (r.room_number, r.check_in_date))

    if all_reservations:
        elements.append(Paragraph('<b>Booking Information</b>', styles['Heading2']))
        booking_data = [['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']]
        for res in all_reservations:
            booking_data.append([
                f"Room {res.room_number}",
                res.voucher_number or '-',
                res.customer_name,
                res.check_in_date.strftime('%y/%m/%d'),
                res.check_out_date.strftime('%y/%m/%d')
            ])
        col_widths = [1.2*inch, 1.8*inch, 2.4*inch, 1.4*inch, 1.4*inch]
        table = Table(booking_data, colWidths=col_widths)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0F0F0F')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(table)
        elements.append(Spacer(1, 0.3*inch))

    summary_data = [
        ['Total Records', str(total)],
    ]
    table = Table(summary_data, colWidths=[3*inch, 2*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C77A1A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)

    doc.build(elements)
    return response


def generate_backtrack_excel_report(date_label, check_ins, check_outs, total, mode):
    """Generate Excel report for backtrack reservations using room_number string field."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from django.http import HttpResponse

    wb = Workbook()
    ws = wb.active
    ws.title = 'Backtrack Report'

    header_fill = PatternFill(start_color='C77A1A', end_color='C77A1A', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=12)

    ws['A1'] = f'Backtrack Report - {date_label}'
    ws['A1'].font = Font(bold=True, size=14, color='C77A1A')
    ws.merge_cells('A1:E1')

    row = 3

    ws[f'A{row}'] = 'Check-ins'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    headers = ['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for res in check_ins:
        ws.cell(row=row, column=1, value=f"Room {res.room_number}")
        ws.cell(row=row, column=2, value=res.voucher_number or '-')
        ws.cell(row=row, column=3, value=res.customer_name)
        ws.cell(row=row, column=4, value=res.check_in_date.strftime('%y/%m/%d'))
        ws.cell(row=row, column=5, value=res.check_out_date.strftime('%y/%m/%d'))
        row += 1

    row += 2

    ws[f'A{row}'] = 'Check-outs'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    headers = ['Room', 'Voucher', 'Customer', 'Check-in', 'Check-out']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    row += 1
    for res in check_outs:
        ws.cell(row=row, column=1, value=f"Room {res.room_number}")
        ws.cell(row=row, column=2, value=res.voucher_number or '-')
        ws.cell(row=row, column=3, value=res.customer_name)
        ws.cell(row=row, column=4, value=res.check_in_date.strftime('%y/%m/%d'))
        ws.cell(row=row, column=5, value=res.check_out_date.strftime('%y/%m/%d'))
        row += 1

    row += 2

    ws[f'A{row}'] = 'Summary'
    ws[f'A{row}'].font = Font(bold=True, size=12)
    row += 1

    ws.cell(row=row, column=1, value='Total Records').font = Font(bold=True)
    ws.cell(row=row, column=2, value=total)

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="backtrack_report_{date_label.replace("/", "-")}.xlsx"'
    wb.save(response)
    return response


@login_required
def backtrack_calendar(request):
    """Room-day matrix showing BacktrackReservation availability per day."""
    today = date.today()
    year_str = request.GET.get('year', '')
    month_str = request.GET.get('month', '')
    try:
        year = int(year_str) if year_str else today.year
        month = int(month_str) if month_str else today.month
        if month < 1 or month > 12:
            raise ValueError
    except ValueError:
        year = today.year
        month = today.month

    num_days = calendar.monthrange(year, month)[1]
    first_day = date(year, month, 1)
    last_day = date(year, month, num_days)

    rooms = Room.objects.filter(is_active=True).annotate(
        room_number_int=Cast('room_number', IntegerField())
    ).order_by('room_number_int')

    reservations = BacktrackReservation.objects.filter(
        status='confirmed',
        check_in_date__lt=last_day + timedelta(days=1),
        check_out_date__gt=first_day,
    )

    # Build lookups: (room_number_str, day_num) -> info
    booked_lookup = {}
    checkout_lookup = {}
    for res in reservations:
        current = res.check_in_date
        while current < res.check_out_date:
            if first_day <= current <= last_day:
                booked_lookup[(res.room_number, current.day)] = {
                    'customer_name': res.customer_name,
                    'voucher_number': res.voucher_number or '',
                    'notes': res.notes or '',
                }
            current += timedelta(days=1)
        checkout_day = res.check_out_date
        if first_day <= checkout_day <= last_day:
            checkout_lookup[(res.room_number, checkout_day.day)] = {
                'customer_name': res.customer_name,
                'voucher_number': res.voucher_number or '',
                'notes': res.notes or '',
            }

    days_list = list(range(1, num_days + 1))
    weekday_abbr = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    day_weekdays = [weekday_abbr[date(year, month, d).weekday()] for d in days_list]

    rooms_data = []
    for room in rooms:
        day_cells = []
        for day_num in days_list:
            d = date(year, month, day_num)
            info = booked_lookup.get((str(room.room_number), day_num))
            checkout_info = checkout_lookup.get((str(room.room_number), day_num))
            is_checkout = checkout_info is not None and info is None
            day_cells.append({
                'day': day_num,
                'is_past': d < today,
                'is_booked': info is not None,
                'is_checkout': is_checkout,
                'customer_name': (checkout_info if is_checkout else info)['customer_name'] if info or is_checkout else '',
                'voucher_number': (checkout_info if is_checkout else info)['voucher_number'] if info or is_checkout else '',
                'notes': (checkout_info if is_checkout else info)['notes'] if info or is_checkout else '',
                'date_display': d.strftime('%a, %b %d, %Y'),
            })
        rooms_data.append({
            'room_number': room.room_number,
            'days': day_cells,
        })

    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    month_names = ['', 'January', 'February', 'March', 'April', 'May', 'June',
                   'July', 'August', 'September', 'October', 'November', 'December']

    context = {
        'year': year,
        'month': month,
        'month_name': month_names[month],
        'prev_month': prev_month,
        'prev_year': prev_year,
        'next_month': next_month,
        'next_year': next_year,
        'rooms_data': rooms_data,
        'days_list': days_list,
        'day_weekdays': day_weekdays,
        'today_day': today.day if today.month == month and today.year == year else None,
    }
    return render(request, 'backtrack/backtrack_calendar.html', context)


@login_required
def backtrack_vouchers_json(request, pk):
    """Return JSON list of vouchers for a backtrack reservation."""
    reservation = get_object_or_404(BacktrackReservation, pk=pk)
    vouchers = reservation.vouchers.all()
    data = [{
        'id': v.id,
        'file_url': v.voucher_file.url,
        'file_name': v.voucher_file.name.split('/')[-1],
        'uploaded_at': v.created_at.strftime('%Y/%m/%d %H:%M'),
        'voucher_number': v.voucher_number or '',
        'customer_name': v.customer_name or '',
    } for v in vouchers]
    return JsonResponse({'vouchers': data})


@login_required
@require_POST
def upload_backtrack_reservation_voucher(request, pk):
    """Upload and link a voucher to a backtrack reservation."""
    reservation = get_object_or_404(BacktrackReservation, pk=pk)
    file = request.FILES.get('voucher_file')
    if not file:
        return JsonResponse({'success': False, 'error': 'No file provided'}, status=400)
    voucher = BacktrackVoucher(voucher_file=file)
    voucher.reservation = reservation
    voucher.is_confirmed = True
    voucher.save()
    return JsonResponse({
        'success': True,
        'voucher': {
            'id': voucher.id,
            'file_url': voucher.voucher_file.url,
            'file_name': voucher.voucher_file.name.split('/')[-1],
            'uploaded_at': voucher.created_at.strftime('%Y/%m/%d %H:%M'),
        }
    })


@login_required
@require_POST
def delete_backtrack_reservation_voucher(request, pk, voucher_id):
    """Delete a voucher linked to a backtrack reservation."""
    voucher = get_object_or_404(BacktrackVoucher, pk=voucher_id, reservation_id=pk)
    voucher.voucher_file.delete(save=False)
    voucher.delete()
    return JsonResponse({'success': True})


@login_required
def backtrack_uploads_page(request, pk):
    """Page showing all vouchers for a backtrack reservation with upload/delete."""
    reservation = get_object_or_404(BacktrackReservation.objects.prefetch_related('vouchers'), pk=pk)
    vouchers = reservation.vouchers.all()

    if request.method == 'POST':
        if 'delete_voucher' in request.POST:
            voucher_id = request.POST.get('voucher_id')
            voucher = get_object_or_404(BacktrackVoucher, pk=voucher_id, reservation_id=pk)
            voucher.voucher_file.delete(save=False)
            voucher.delete()
            messages.success(request, 'Voucher deleted.')
            return redirect('backtrack:backtrack_uploads', pk=pk)

        if 'voucher_file' in request.FILES:
            voucher = BacktrackVoucher(voucher_file=request.FILES['voucher_file'])
            voucher.reservation = reservation
            voucher.is_confirmed = True
            voucher.save()
            messages.success(request, 'Voucher uploaded.')
            return redirect('backtrack:backtrack_uploads', pk=pk)

    return render(request, 'backtrack/backtrack_uploads.html', {
        'reservation': reservation,
        'vouchers': vouchers,
    })
